from bs4 import BeautifulSoup
import urllib.request
import string
from pprint import pprint
import json
from openpyxl import load_workbook, Workbook

from dataclasses import dataclass

@dataclass
class NobelLit:
    date: str
    wiki: str
    name: str
    country: str
    genre: str
    width: str
    height: str
    source: str



def write_result(result):

    keys = []
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sname='maleoscar'
    try:
        del workbook[sname]
    except:
        pass

    ws=workbook.create_sheet(sname)

    for i in range(len(result)) :
        sub_obj = result[i].__dict__
        if i == 0 :
            keys = list(sub_obj.keys())
            for k in range(len(keys)) :
                # row or column index start from 1
                ws.cell(row = (i + 1), column = (k + 1), value = keys[k]);
        for j in range(len(keys)) :
            ws.cell(row = (i + 2), column = (j + 1), value = sub_obj[keys[j]]);
    workbook.save(filename="../monarchs.xlsx") #"new-users.xlsx")





def get_oscar():
    pp=None
    r=None
    def ev(x,pp,r):
        try:
            q=eval(x)
        except:
            q='ERROR'
        return q
    
    mapping=Oscar(
        "pp.a['href']",
        "pp.a.text",
        "pp.parent.th.a.text",
        "pp.parent.findAll('td')[2].a.text",
        "pp.parent.findAll('td')[1].b.text",
        "r['description']",
        "r['thumbnail']['width']",
        "r['thumbnail']['height']",
        "r['thumbnail']['source']",
        )
    
    
    result=[]
    fp=urllib.request.urlopen("https://en.wikipedia.org/wiki/Academy_Award_for_Best_Actress")
    soup = BeautifulSoup(fp, 'html.parser')
    #s=soup.findAll(name='table',attrs={'class':'wikitable unsortable'})
    x=soup.findAll('img',attrs={'alt':'Award winner'})

    for q in x[1:]:
        #print(q.parent)
        pp=q.parent.parent
        href=eval("pp.a['href']")
        url="https://en.wikipedia.org/api/rest_v1/page/summary/"+href[6:]
        with urllib.request.urlopen(url) as fp:
            buff=fp.read().decode('utf-8')
            r=json.loads(buff)
            osc=Oscar(*map(lambda x:ev(x,pp,r),mapping.__dict__.values()))
        
        result.append(osc)
    

        """
        try:
            try:
                a=pp.a["href"]
            except:
                a='ERROR'
            try:
                b=pp.a.text
            except:
                b='ERROR'
            try:
                c=pp.parent.th.a.text
            except:
                c='ERROR'
            try:
                d=pp.parent.findAll('td')[2].a.text
            except:
                d="ERROR"
            try:
                e=pp.parent.findAll('td')[1].b.text
            except:
                e="ERROR"
                
            w2=w[6:]
            with urllib.request.urlopen("https://en.wikipedia.org/api/rest_v1/page/summary/"+w2) as fp:
                buff=fp.read().decode('utf-8')
                r=json.loads(buff)
                try:
                    sheet[f'F{i}']=r['description']
                    sheet[f'G{i}']=r['thumbnail']['width']
                    sheet[f'H{i}']=r['thumbnail']['height']
                    sheet[f'I{i}']=r['thumbnail']['source']
                except KeyError:
                    pass
            osc=Oscar(a,b,c,d,e)
            result.append(osc)
        except Exception as e:
            print("error")
            """
    return result
    i=1

    # q.parent.parent.parent.th.a.text
    #q.parent.parent.a.text
    #q.parent.parent.a['href']

def getThumbnails():
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sheet=workbook['maleoscar']



    result=[]
    for i in range(1,len(list(sheet.rows))+1):
        w=sheet.cell(i,1).value
        w2=w[6:]
        with urllib.request.urlopen("https://en.wikipedia.org/api/rest_v1/page/summary/"+w2) as fp:
            buff=fp.read().decode('utf-8')
            r=json.loads(buff)
            try:
                sheet[f'E{i}']=r['description']
                sheet[f'F{i}']=r['thumbnail']['width']
                sheet[f'G{i}']=r['thumbnail']['height']
                sheet[f'H{i}']=r['thumbnail']['source']
            except KeyError:
                pass






def get_lit():
    result=[]
    pp=None
    r=None
    def either(a,b):
        try:
            q=eval(a)
        except:
            try:
                q=eval(b)
            except:
                q="ERROR"
        return q
    
    def ev(x,pp,r):
        try:
            q=eval(x)
        except:
            q='ERROR'
        return q
    
    mapping=NobelLit(
        #"pp[1].a['href']",
        "pp[0].text",
        "pp[2].a['href']",
        "pp[2].a.text",
        "pp[3].a.text",
        "pp[6].text",
        "r['thumbnail']['width']",
        "r['thumbnail']['height']",
        "r['thumbnail']['source']",
        )
    
    
    fp=urllib.request.urlopen("https://en.wikipedia.org/wiki/List_of_Nobel_laureates_in_Literature")
    soup = BeautifulSoup(fp, 'html.parser')
    s=soup.findAll(name='table',attrs={'class':'wikitable sortable'})
    
    t2=s[0].findAll(name="tbody")
    
    tr=t2[0].findAll(name="tr")
    
    for ll in tr[1:]:
        try:
            
            pp=ll.findAll('td')
            href=eval("pp[2].a['href']")
            url="https://en.wikipedia.org/api/rest_v1/page/summary/"+href[6:]
            with urllib.request.urlopen(url) as fp:
                buff=fp.read().decode('utf-8')
                r=json.loads(buff)
                c=NobelLit(*map(lambda x:ev(x,pp,r),mapping.__dict__.values()))
                print(c)
                result.append(c)
        except:
            print(pp)
    return result

result=get_lit()        
#result=[Oscar(wiki='ERROR', name='ERROR', date='ERROR'), Oscar(wiki='/wiki/Emil_Jannings', name='Emil Jannings', date='Emil Jannings'), Oscar(wiki='/wiki/Warner_Baxter', name='Warner Baxter', date='Warner Baxter')]
pprint(result)
write_result(result)
i=1

pprint(result)
#getThumbnails()    