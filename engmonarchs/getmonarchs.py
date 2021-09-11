from bs4 import BeautifulSoup
import urllib.request
import string
from pprint import pprint
import json
from openpyxl import load_workbook

def getMonarchsFromWiki():
        
    with urllib.request.urlopen("https://en.wikipedia.org/wiki/Yale_English_Monarchs_series") as fp:
        soup = BeautifulSoup(fp, 'html.parser')
        tr=soup.table.find_all('tr')
        for r in range(1,len(tr)):
            td=tr[r].find_all('td')
            aref=td[0].find('a')
            at=aref.attrs
            dates=td[1].string
            if not dates:
                dates="unknown" 
            print(f'{at["href"]}|{at["title"]}|{aref.string}|{str.strip(dates)}')


def getThumbnails():
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sheet=workbook['monarchs']
    
    
    
    result=[]
    for i in range(1,len(list(sheet.rows))+1):
        w=sheet.cell(i,1).value
        w2=w[6:]
        with urllib.request.urlopen("https://en.wikipedia.org/api/rest_v1/page/summary/"+w2) as fp:
            buff=fp.read().decode('utf-8')
            r=json.loads(buff)
            try:
                sheet[f'E{i}']=r['description']
                sheet[f'F{i}']=r['thumbnail']['width']
                sheet[f'G{i}']=r['thumbnail']['height']
                sheet[f'H{i}']=r['thumbnail']['source']
            except KeyError:
                pass
        
        

    workbook.save(filename="../mon2.xlsx")
    
def getSummary():
    def getv(v):
        return getattr(v,'value').decode('utf-8')
    
    
    
    result=[]
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sheet=workbook['monarchs']
    for i,row in enumerate(sheet.rows):
        rv=map(getv,row)
        if i==0:
            head=list(rv)
        else:
            result.append(dict(zip(head,rv)))
    pprint(result)
    i=1
    
    
    
getSummary()    





    
"""
https://en.wikipedia.org/api/rest_v1/page/summary/William_II_of_England
{"type":"standard","title":"William II of England","displaytitle":"William II of England","namespace":{"id":0,"text":""},"wikibase_item":"Q102005","titles":{"canonical":"William_II_of_England","normalized":"William II of England","display":"William II of England"},"pageid":33918,"thumbnail":{"source":"https://upload.wikimedia.org/wikipedia/commons/thumb/9/92/William_II_of_England.jpg/192px-William_II_of_England.jpg","width":192,"height":320},"originalimage":{"source":"https://upload.wikimedia.org/wikipedia/commons/9/92/William_II_of_England.jpg","width":304,"height":506},"lang":"en","dir":"ltr","revision":"1036904980","tid":"6667bc30-f9ea-11eb-ba3c-338d0699ca8e","timestamp":"2021-08-03T11:23:55Z","description":"King of England","description_source":"local","content_urls":{"desktop":{"page":"https://en.wikipedia.org/wiki/William_II_of_England","revisions":"https://en.wikipedia.org/wiki/William_II_of_England?action=history","edit":"https://en.wikipedia.org/wiki/William_II_of_England?action=edit","talk":"https://en.wikipedia.org/wiki/Talk:William_II_of_England"},"mobile":{"page":"https://en.m.wikipedia.org/wiki/William_II_of_England","revisions":"https://en.m.wikipedia.org/wiki/Special:History/William_II_of_England","edit":"https://en.m.wikipedia.org/wiki/William_II_of_England?action=edit","talk":"https://en.m.wikipedia.org/wiki/Talk:William_II_of_England"}},"extract":"William II, the third son of William the Conqueror, was King of England from 26 September 1087 until his death in 1100, with powers over Normandy, and influence in Scotland. He was less successful in extending control into Wales. William is commonly referred to as William Rufus, perhaps because of his ruddy appearance or, more likely, due to having red hair as a child that grew out in later life.","extract_html":"<p><b>William II</b>, the third son of William the Conqueror, was King of England from 26 September 1087 until his death in 1100, with powers over Normandy, and influence in Scotland. He was less successful in extending control into Wales. William is commonly referred to as <b>William Rufus</b>, perhaps because of his ruddy appearance or, more likely, due to having red hair as a child that grew out in later life.</p>"}
"""    
    
i=1