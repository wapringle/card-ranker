from bs4 import BeautifulSoup
import urllib.request
import string
from pprint import pprint
import json
from openpyxl import load_workbook

def getPresidentsFromWiki():
        
    with urllib.request.urlopen("https://en.wikipedia.org/wiki/List_of_presidents_of_the_United_States") as fp:
        soup = BeautifulSoup(fp, 'html.parser')
        
        i=1
        
        tt=soup.findAll('table')[1]
        rr=tt.findAll('tr')
        
        for q in rr:
            td=q.findAll('td')
            if len(td)>2:
                q=td[2].a
                print(f'{q["href"]}|{q.text}|{td[0].text[:-1]}')
                
        i=1



"""
<a href="/wiki/George_Washington" title="George Washington">George Washington</a>
<a href="/wiki/John_Adams" title="John Adams">John Adams</a>
<a href="/wiki/Thomas_Jefferson" title="Thomas Jefferson">Thomas Jefferson</a>
<a href="/wiki/James_Madison" title="James Madison">James Madison</a>
<a href="/wiki/James_Monroe" title="James Monroe">James Monroe</a>
<a href="/wiki/John_Quincy_Adams" title="John Quincy Adams">John Quincy Adams</a>
<a href="/wiki/Andrew_Jackson" title="Andrew Jackson">Andrew Jackson</a>
<a href="/wiki/Martin_Van_Buren" title="Martin Van Buren">Martin Van Buren</a>
<a href="/wiki/William_Henry_Harrison" title="William Henry Harrison">William Henry Harrison</a>
<a href="/wiki/John_Tyler" title="John Tyler">John Tyler</a>
<a href="/wiki/James_K._Polk" title="James K. Polk">James K. Polk</a>
<a href="/wiki/Zachary_Taylor" title="Zachary Taylor">Zachary Taylor</a>
<a href="/wiki/Millard_Fillmore" title="Millard Fillmore">Millard Fillmore</a>
<a href="/wiki/Franklin_Pierce" title="Franklin Pierce">Franklin Pierce</a>
<a href="/wiki/James_Buchanan" title="James Buchanan">James Buchanan</a>
<a href="/wiki/Abraham_Lincoln" title="Abraham Lincoln">Abraham Lincoln</a>
<a href="/wiki/1864_United_States_presidential_election" title="1864 United States presidential election">1864</a>
<a href="/wiki/Andrew_Johnson" title="Andrew Johnson">Andrew Johnson</a>
<a href="/wiki/Ulysses_S._Grant" title="Ulysses S. Grant">Ulysses S. Grant</a>
<a href="/wiki/Rutherford_B._Hayes" title="Rutherford B. Hayes">Rutherford B. Hayes</a>
<a href="/wiki/James_A._Garfield" title="James A. Garfield">James A. Garfield</a>
<a href="/wiki/Chester_A._Arthur" title="Chester A. Arthur">Chester A. Arthur</a>
<a href="/wiki/Grover_Cleveland" title="Grover Cleveland">Grover Cleveland</a>
<a href="/wiki/Benjamin_Harrison" title="Benjamin Harrison">Benjamin Harrison</a>
<a href="/wiki/Grover_Cleveland" title="Grover Cleveland">Grover Cleveland</a>
<a href="/wiki/William_McKinley" title="William McKinley">William McKinley</a>
<a href="/wiki/Theodore_Roosevelt" title="Theodore Roosevelt">Theodore Roosevelt</a>
<a href="/wiki/William_Howard_Taft" title="William Howard Taft">William Howard Taft</a>
<a href="/wiki/Woodrow_Wilson" title="Woodrow Wilson">Woodrow Wilson</a>
<a href="/wiki/Warren_G._Harding" title="Warren G. Harding">Warren G. Harding</a>
<a href="/wiki/Calvin_Coolidge" title="Calvin Coolidge">Calvin Coolidge</a>
<a href="/wiki/Herbert_Hoover" title="Herbert Hoover">Herbert Hoover</a>
<a href="/wiki/Franklin_D._Roosevelt" title="Franklin D. Roosevelt">Franklin D. Roosevelt</a>
<a href="/wiki/Harry_S._Truman" title="Harry S. Truman">Harry S. Truman</a>
<a href="/wiki/Dwight_D._Eisenhower" title="Dwight D. Eisenhower">Dwight D. Eisenhower</a>
<a href="/wiki/John_F._Kennedy" title="John F. Kennedy">John F. Kennedy</a>
<a href="/wiki/Lyndon_B._Johnson" title="Lyndon B. Johnson">Lyndon B. Johnson</a>
<a href="/wiki/Richard_Nixon" title="Richard Nixon">Richard Nixon</a>
<a href="/wiki/Gerald_Ford" title="Gerald Ford">Gerald Ford</a>
<a href="/wiki/Jimmy_Carter" title="Jimmy Carter">Jimmy Carter</a>
<a href="/wiki/Ronald_Reagan" title="Ronald Reagan">Ronald Reagan</a>
<a href="/wiki/George_H._W._Bush" title="George H. W. Bush">George H. W. Bush</a>
<a href="/wiki/Bill_Clinton" title="Bill Clinton">Bill Clinton</a>
<a href="/wiki/George_W._Bush" title="George W. Bush">George W. Bush</a>
<a href="/wiki/Barack_Obama" title="Barack Obama">Barack Obama</a>
<a href="/wiki/Donald_Trump" title="Donald Trump">Donald Trump</a>
<a href="/wiki/Joe_Biden" title="Joe Biden">Joe Biden</a>
"""

# https://en.wikipedia.org/api/rest_v1/page/summary/Harry_S._Truman
def getThumbnails():
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sheet=workbook['presidents']
    
    
    
    result=[]
    for i in range(1,len(list(sheet.rows))+1):
        w=sheet.cell(i,1).value
        w2=w[6:]
        with urllib.request.urlopen("https://en.wikipedia.org/api/rest_v1/page/summary/"+w2) as fp:
            buff=fp.read().decode('utf-8')
            r=json.loads(buff)
            try:
                sheet[f'E{i}']=r['description']
                sheet[f'F{i}']=r['thumbnail']['width']
                sheet[f'G{i}']=r['thumbnail']['height']
                sheet[f'H{i}']=r['thumbnail']['source']
            except KeyError:
                pass
        
        

    workbook.save(filename="../mon2.xlsx")
    
def getSummary():
    def getv(v):
        return getattr(v,'value')#.decode('utf-8')
    
    
    
    result=[]
    workbook = load_workbook(filename="../monarchs.xlsx")#, read_only=True, data_only=True)
    sheet=workbook['presidents']
    for i,row in enumerate(sheet.rows):
        rv=map(getv,row)
        if i==0:
            head=list(rv)
        else:
            dd=dict(zip(head,rv))
            print(dd)
            result.append(dict(dd))
    pprint(result)
    i=1
    
    
#getPresidentsFromWiki()  
getSummary()    
i=1





    
"""
https://en.wikipedia.org/api/rest_v1/page/summary/William_II_of_England
{"type":"standard","title":"William II of England","displaytitle":"William II of England","namespace":{"id":0,"text":""},"wikibase_item":"Q102005","titles":{"canonical":"William_II_of_England","normalized":"William II of England","display":"William II of England"},"pageid":33918,"thumbnail":{"source":"https://upload.wikimedia.org/wikipedia/commons/thumb/9/92/William_II_of_England.jpg/192px-William_II_of_England.jpg","width":192,"height":320},"originalimage":{"source":"https://upload.wikimedia.org/wikipedia/commons/9/92/William_II_of_England.jpg","width":304,"height":506},"lang":"en","dir":"ltr","revision":"1036904980","tid":"6667bc30-f9ea-11eb-ba3c-338d0699ca8e","timestamp":"2021-08-03T11:23:55Z","description":"King of England","description_source":"local","content_urls":{"desktop":{"page":"https://en.wikipedia.org/wiki/William_II_of_England","revisions":"https://en.wikipedia.org/wiki/William_II_of_England?action=history","edit":"https://en.wikipedia.org/wiki/William_II_of_England?action=edit","talk":"https://en.wikipedia.org/wiki/Talk:William_II_of_England"},"mobile":{"page":"https://en.m.wikipedia.org/wiki/William_II_of_England","revisions":"https://en.m.wikipedia.org/wiki/Special:History/William_II_of_England","edit":"https://en.m.wikipedia.org/wiki/William_II_of_England?action=edit","talk":"https://en.m.wikipedia.org/wiki/Talk:William_II_of_England"}},"extract":"William II, the third son of William the Conqueror, was King of England from 26 September 1087 until his death in 1100, with powers over Normandy, and influence in Scotland. He was less successful in extending control into Wales. William is commonly referred to as William Rufus, perhaps because of his ruddy appearance or, more likely, due to having red hair as a child that grew out in later life.","extract_html":"<p><b>William II</b>, the third son of William the Conqueror, was King of England from 26 September 1087 until his death in 1100, with powers over Normandy, and influence in Scotland. He was less successful in extending control into Wales. William is commonly referred to as <b>William Rufus</b>, perhaps because of his ruddy appearance or, more likely, due to having red hair as a child that grew out in later life.</p>"}
"""    
    
i=1